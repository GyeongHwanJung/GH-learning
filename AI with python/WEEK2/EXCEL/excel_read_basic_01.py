import openpyxl

filename="../Data/sample.xlsx"

book=openpyxl.load_workbook(filename)

sheet=book.worksheets[0]
print('sheet =>', sheet.title)
print('sheet.max_row =>', sheet.max_row)
print('sheet.max_column =>', sheet.max_column)

row_data=[]
for row in sheet.rows:
    row_data.append([row[0].value, row[sheet.max_column-1].value])
print('len(row_data) => ', len(row_data), row_data)