from openpyxl import Workbook

filename='../Data/sample.xlsx'

wb=Workbook()
ws=wb.active
ws.title='new sheet'
ws['A1']='Language'
ws['B1']='Create'

wb.save(filename=filename)